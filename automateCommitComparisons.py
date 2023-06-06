import pandas as pd
import os
import openpyxl

# Point path to folder containing Excel files
input_folder = "/Users/dianacrisan/Desktop/Dizertatie/ref-miner/DesigniteJava/CommitsResults"

# Combine the output CSV files into a single Excel file
output_folder = "/Users/dianacrisan/Desktop/Dizertatie/ref-miner/DesigniteJava/EvaluationCommitsResults"

# Get a list of all Excel files in the folder, sorted by creation date
file_list = sorted([f for f in os.listdir(input_folder) if f.endswith('.xlsx')], key=lambda f: os.stat(os.path.join(input_folder, f)).st_mtime)

# Create a new workbook to store the differences
diff_workbook = openpyxl.Workbook()

# Delete the first empty sheet
diff_workbook.remove(diff_workbook.active)

# Loop through each pair of files
for i in range(len(file_list)-1):
    # Save commit IDs from file names
    first_commit_id = (file_list[i])[7:-5]
    second_commit_id = (file_list[i+1])[7:-5]

    # Load the two input workbooks
    workbook1_path = os.path.join(input_folder, file_list[i])
    workbook2_path = os.path.join(input_folder, file_list[i+1])
    workbook1 = openpyxl.load_workbook(workbook1_path)
    workbook2 = openpyxl.load_workbook(workbook2_path)

    # Loop through each sheet in workbook1 and compare to the corresponding sheet in workbook2
    for sheet_name in workbook1.sheetnames:
        worksheet1 = workbook1[sheet_name]
        worksheet2 = workbook2[sheet_name]

        # Create a new sheet in the differences workbook to store the differences for this pair of sheets
        diff_sheet = diff_workbook.create_sheet(sheet_name + "_diff")

        # Combine the two worksheets into one list of rows
        rows = list(worksheet1.iter_rows(values_only=True)) + list(worksheet2.iter_rows(values_only=True))

        # Find the unique rows in each workbook
        workbook1_rows = set(worksheet1.iter_rows(values_only=True))
        workbook2_rows = set(worksheet2.iter_rows(values_only=True))
        unique_rows_in_workbook1 = workbook1_rows - workbook2_rows
        unique_rows_in_workbook2 = workbook2_rows - workbook1_rows

        # Add the unique rows to the diff sheet and color them accordingly
        for row_num, row in enumerate(rows):
            if row in unique_rows_in_workbook1:
                for col_num, cell in enumerate(row):
                    diff_sheet.cell(row=row_num+1, column=col_num+1).value = cell
                    diff_sheet.cell(row=row_num+1, column=col_num+1).fill = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif row in unique_rows_in_workbook2:
                for col_num, cell in enumerate(row):
                    diff_sheet.cell(row=row_num+1, column=col_num+1+len(row)).value = cell
                    diff_sheet.cell(row=row_num+1, column=col_num+1+len(row)).fill = openpyxl.styles.PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # Save the differences workbook to a new Excel file
    output_filename = os.path.join(output_folder, 'diff-' + first_commit_id + '-' + second_commit_id + '.xlsx')
    diff_workbook.save(output_filename)
